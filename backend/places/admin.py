import datetime

import openpyxl
import xlsxwriter
from django.contrib import admin, messages
from django.contrib.gis.admin import GISModelAdmin
from django.contrib.gis.geos import Point
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path

from .models import Place, WeatherSummary


@admin.register(Place)
class PlaceAdmin(GISModelAdmin):
    list_display = ("name", "rating")
    search_fields = ("name",)
    change_list_template = "admin/places_change_list.html"
    actions = ("export_xlsx",)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-xlsx/",
                self.admin_site.admin_view(self.import_xlsx),
                name="import-xlsx",
            ),
        ]
        return custom_urls + urls

    def import_xlsx(self, request):
        if request.method == "POST" and request.FILES.get("xlsx_file"):
            xlsx_file = request.FILES["xlsx_file"]
            try:
                workbook = openpyxl.load_workbook(xlsx_file)
                sheet = workbook.active
                imported_count = 0

                existing_keys = {
                    (place.name, place.rating, place.location.wkt)
                    for place in Place.objects.all()
                }

                existing_keys = {
                    (place.name, place.rating, place.location.wkt)
                    for place in Place.objects.all()
                }

                new_places = []
                imported_count = 0

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name, coord_str, rating = row

                    if None in (name, coord_str, rating):
                        continue

                    if not (0 <= rating <= 25):
                        raise ValidationError(
                            f"Invalid rating {rating}. Must be between 0 and 25."
                        )

                    try:
                        latitude, longitude = map(float, coord_str.split(","))
                    except ValueError:
                        raise ValidationError(f"Invalid coordinate format: {coord_str}")

                    point = Point(longitude, latitude)
                    key = (name, rating, point.wkt)
                    if key in existing_keys:
                        continue

                    new_places.append(Place(name=name, location=point, rating=rating))
                    imported_count += 1
                    existing_keys.add(key)

                if new_places:
                    Place.objects.bulk_create(new_places)

                self.message_user(
                    request, f"Imported {imported_count} locations.", messages.SUCCESS
                )
                return redirect("..")

            except Exception as e:
                self.message_user(request, f"Error: {str(e)}", messages.ERROR)

        return render(request, "admin/import_xlsx.html", {})

    def import_xlsx_button(self):
        return '<a href="import-xlsx/" class="button">Import data from XLSX</a>'

    @admin.action(description="Export selected places to XLSX")
    def export_xlsx(self, request, queryset):
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = (
            f"places_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"

        workbook = xlsxwriter.Workbook(response, {"in_memory": True})
        worksheet = workbook.add_worksheet("Places")

        headers = ("Name", "Coordinates", "Rating")
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header)

        row = 1
        for place in queryset:
            worksheet.write(row, 0, place.name)
            worksheet.write(row, 1, f"{place.location.x}, {place.location.y}")
            worksheet.write(row, 2, place.rating)
            row += 1

        workbook.close()
        return response


@admin.register(WeatherSummary)
class WeatherSummaryAdmin(admin.ModelAdmin):
    list_display = (
        "place",
        "timestamp",
        "temperature",
        "humidity",
        "pressure",
        "wind_direction",
        "wind_speed",
    )
    list_filter = ("place", "timestamp")
    date_hierarchy = "timestamp"
    actions = ["export_to_xlsx"]
    readonly_fields = [field.name for field in WeatherSummary._meta.fields]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.action(description="Export selected weather summaries to XLSX")
    def export_to_xlsx(self, request, queryset):
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"weather_summary_{
            datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        }.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"

        workbook = xlsxwriter.Workbook(response, {"in_memory": True})
        worksheet = workbook.add_worksheet("Weather report")

        headers = [
            "Place",
            "Time",
            "Temperature (°C)",
            "Humidity (%)",
            "Pressure (mmHg)",
            "Wind direction",
            "Wind speed (m/s)",
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header)

        row = 1
        for weather in queryset:
            worksheet.write(row, 0, weather.place.name)
            worksheet.write(row, 1, weather.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            worksheet.write(row, 2, weather.temperature)
            worksheet.write(row, 3, weather.humidity)
            worksheet.write(row, 4, weather.pressure)
            worksheet.write(row, 5, weather.wind_direction)
            worksheet.write(row, 6, weather.wind_speed)
            row += 1

        workbook.close()
        return response
